import telebot
import os
import json 
import openpyxl
import config 
import timer
from openpyxl import load_workbook
from telebot.types import Message
from telebot import types



bot = telebot.TeleBot(config.TOKEN)

global BUFFER
BUFFER = {}

# Меню регистрации
def registration(message):
    BUFFER = {}
    keyboard = types.InlineKeyboardMarkup()
    button_student = 'Студент'
    button_visitor = 'Посетитель'

    keyboard.add(types.InlineKeyboardButton(text = button_student, callback_data = 'student'))
    keyboard.add(types.InlineKeyboardButton(text = button_visitor, callback_data = 'visitor'))

    bot.send_message(message.chat.id,'Зарегистрировться как:', reply_markup = keyboard)

# /start
@bot.message_handler(commands=['start'])
def start_message(message:Message):
    print ('Нажили start', message)
    bot.send_message(message.chat.id,'Привет, я IRNITU_bot!\n' + 'Проидите регистрацию')
    registration(message)

@bot.message_handler(commands=['reg'])
def repeat_registration(message):
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    btn1 = types.KeyboardButton('Продолжить')
    btn2 = types.KeyboardButton('Отмена')
    markup.add(btn1, btn2)

    msg = bot.send_message(message.chat.id, 'Вы уверены, что хотите пройти повторную регистрацию?\n' + 
                     'После нажатия кнопки Продолжить, отмена будет невозможна!' +
                     'Текущая ученная запись будет удалена', reply_markup = markup)
    bot.register_next_step_handler(msg, repeat_registration_answer)


def repeat_registration_answer(message):
    if message.text == 'Продолжить':
        print('Продолжить')
        content = read()
        del content[str(message.chat.id)]
        save(content)
        registration(message)

    else:
        if not str(message.chat.id) in read():
            bot.send_message(message.chat.id, 'Вы ещё не зарегистрировались!')
            return registration(message)
            
        user_status = read()[str(message.chat.id)]['status']
        if user_status == 'student':
            return bot.send_message(message.chat.id, 'Возвращено в основное меню', reply_markup = keyboard_main_menu_visitor())
        elif user_status == 'visitor':
            return bot.send_message(message.chat.id, 'Возвращено в основное меню', reply_markup = keyboard_main_menu_visitor())
        elif user_status == 'admin':
            pass

 



# Обработка ответов от чат-клавиатуры
@bot.callback_query_handler(func=lambda message:True)
def ans(message:Message):
    print ('Кнопка', message.message.json)
    chat_id = message.message.chat.id
    global BUFFER
    # регистрация студента
    if message.data == 'student':
        if str(chat_id) in read() and not check_start_reg(chat_id):
            bot.send_message(chat_id,'Вы уже зарегистрированы!')
            return

        if not check_start_reg(chat_id, True):
            BUFFER[chat_id] = 'студент'
            msg = bot.send_message(chat_id,'Введите ФИО (через пробел)')
            bot.register_next_step_handler(msg, ask_name)

    # Регистрация посетителя
    if message.data == 'visitor':
        if str(chat_id) in read() and not check_start_reg(chat_id):
            bot.send_message(chat_id,'Вы уже зарегистрированы!')
            return

        if not check_start_reg(chat_id, True):
            BUFFER[chat_id] = 'посетитель'
            msg = bot.send_message(chat_id, 'Введите номер договора')
            bot.register_next_step_handler(msg, ask_contract)

    # Описание оборудования
    if message.data[:7] == 'info_eq':
        name =  message.data[7:-1]
        keyboard = types.InlineKeyboardMarkup()
        button = 'Свернуть описание'
        keyboard.add(types.InlineKeyboardButton(text = button, callback_data = 'close_info_eq' + name))

        bot.edit_message_text(name + info_equipment(name, 'Оборудование'), 
                              chat_id, message.message.message_id, reply_markup = keyboard)

    # Свернуть описание обоудования
    if message.data[:13] == 'close_info_eq':
        name = message.data[13:]   

        keyboard = types.InlineKeyboardMarkup()
        button = 'Описание'
        keyboard.add(types.InlineKeyboardButton(text = button, callback_data = 'info_eq' + name + '\n'))
                                                   
        bot.edit_message_text(name, chat_id, message.message.message_id, reply_markup = keyboard)

    # Кнопка израсходовать
    if message.data[:13] == 'spend_exp_mat':
        name = message.data[13:]
        if str(chat_id) in BUFFER and BUFFER[str(chat_id)][-1] == 's' :
            
            bot.send_message(chat_id, f'Вы уже нажали на кнопку Израсходовать  ({BUFFER[str(chat_id)][:-1]})'+
                            '\nВведите количество, либо нажмите кнопку Отмена')

        elif str(chat_id) in BUFFER and BUFFER[str(chat_id)][-1] == 'b' :
            
            bot.send_message(chat_id, f'Вы уже нажали на кнопку Вернуть  ({BUFFER[str(chat_id)][:-1]})'+
                            '\nВведите количество, либо нажмите кнопку Отмена')
        else:
            
            markup = types.ReplyKeyboardMarkup(one_time_keyboard=False, resize_keyboard=True)
            btn = types.KeyboardButton('Отмена')
            markup.add(btn)

            BUFFER[str(chat_id)] = name + 's' # Записываем в буфер название материала который выбрал пользователь и действие (s-spend)
            BUFFER['m_id' + str(chat_id)] = message.message.message_id #записываем в буфер id сообщения для дальнейшего редактирования
            
            msg = bot.send_message(chat_id, name +'\nВведите количество, которое хотите израсходовать', reply_markup = markup)
            bot.register_next_step_handler(msg, change_kol)

    # Кнопка вернуть 
    if message.data[:20] == 'cancel_spend_exp_mat':
        name = message.data[20:]
        
        if str(chat_id) in BUFFER and BUFFER[str(chat_id)][-1] == 'b':

            bot.send_message(chat_id, f'Вы уже нажали на кнопку Вернуть ({BUFFER[str(chat_id)][:-1]})'+
                            '\nВведите количество, либо нажмите кнопку Отмена')

        elif str(chat_id) in BUFFER and BUFFER[str(chat_id)][-1] == 's' :
            
            bot.send_message(chat_id, f'Вы уже нажали на кнопку Израсходовать  ({BUFFER[str(chat_id)][:-1]})'+
                            '\nВведите количество, либо нажмите кнопку Отмена')

        else:
            
            markup = types.ReplyKeyboardMarkup(one_time_keyboard=False, resize_keyboard=True)
            btn = types.KeyboardButton('Отмена')
            markup.add(btn)

            BUFFER[str(chat_id)] = name + 'b' # Записываем в буфер название материала который выбрал пользователь и действие (b-back)
            BUFFER['m_id' + str(chat_id)] = message.message.message_id # Записываем в буфер id сообщения для дальнейшего редактирования
            
            msg = bot.send_message(chat_id, name +'\nВведите количество, которое хотите вернуть', reply_markup = markup)
            bot.register_next_step_handler(msg, change_kol)



#======================================================Обработка текстовых сообщений=================================#
@bot.message_handler(content_types=['text'])
def text(message:Message):
    print ('Что-то написали', message.json)

    chat_id = message.chat.id

    # Переход в режим Администратора
    if message.text == config.ADMIN_PASS:
        add_user(chat_id, key='admin')
        bot.send_message(chat_id, 'Вы успешно зарегистрировались как Администратор!', reply_markup = keyboard_main_menu_admin())


    # Если пользователь что-то написал до регистрации
    if not str(chat_id) in read():
        bot.send_message(message.chat.id, 'Вы ещё не зарегистрировались!')
        registration(message)
        return


    # Присваивание статуса пользователю
    user_status = read()[str(chat_id)]['status']


#===========================ДЛЯ АДМИНИСТРАТОРА==============================#
    if user_status == 'admin':
        if message.text == 'Основное меню':
            bot.send_message(chat_id, 'Возвращено в основное меню', reply_markup = keyboard_main_menu_admin())
        if message.text == 'Написать сообщение посетителям':
            markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
            btn1 = types.KeyboardButton('Отмена')
            markup.add(btn1)
            msg = bot.send_message(chat_id, 'Введите сообщение для отправки (после того как вы отпраите сообщение в чат, оно придёт всем посетителям)',
                                   reply_markup=markup)
            bot.register_next_step_handler(msg, send_message_visitors)

        if message.text == 'Написать сообщение студентам':
            markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
            btn1 = types.KeyboardButton('Отмена')
            markup.add(btn1)
            msg = bot.send_message(chat_id, 'Введите сообщение для отправки (после того как вы отпраите сообщение в чат, оно придёт всем посетителям)',
                                   reply_markup=markup)
            bot.register_next_step_handler(msg, send_message_students)

#===========================================================================#


#===========================ДЛЯ ПОСЕТИТЕЛЕЙ=================================#
    
    if user_status == 'visitor':
        # Расписание
        if message.text == 'Расписание занятий':
            bot.send_message(chat_id, timetable_visitor(chat_id))
    
        # Отработка
        elif message.text == 'Ближайшая отработка':
            bot.send_message(chat_id, otrabotka(chat_id))

        # Задолжность по оплате
        elif message.text == 'Задолжность по оплате':
            bot.send_message(chat_id, owe(chat_id))


            # Основное меню для посетителей
        elif message.text == 'Основное меню':
            bot.send_message(chat_id, 'Возвращено в основное меню', reply_markup = keyboard_main_menu_visitor())
        else:
            bot.send_message(chat_id, 'Я вас не понимаю')
            bot.send_message(chat_id, 'Выберите пункт из меню', reply_markup = keyboard_main_menu_visitor())

#===========================ДЛЯ СТУДЕНТОВ=================================#
        
    if user_status == 'student':
        # Вывод списка оборудования
        if message.text == 'Оборудование' :
            eq = change_BD('Оборудование')
            bot.send_message(message.chat.id, 'Список оборудования:', reply_markup = main_menu_student())
            for i in range(len(eq)):
                button = 'Описание'
                keyboard = types.InlineKeyboardMarkup()
                keyboard.add(types.InlineKeyboardButton(text = button, callback_data = 'info_eq' + 
                                                        change_BD('Оборудование')[i]))
                bot.send_message(message.chat.id, change_BD('Оборудование')[i], reply_markup = keyboard)    
        # Вывод списка инструментов
        elif message.text == 'Инструмент':
            tools = change_BD('Инструмент')
            tools_str =''
            bot.send_message(message.chat.id, 'Доступный инструмент:', reply_markup = main_menu_student())
            for i in tools:
                tools_str = tools_str + i
            bot.send_message(message.chat.id, tools_str)

        # Вывод списка расходных материалов
        elif message.text == 'Расходные материалы':
            markup = types.ReplyKeyboardMarkup(one_time_keyboard=False, resize_keyboard=True)
            btn1 = types.KeyboardButton('Израсходовать')
            btn2 = types.KeyboardButton('Вернуть')
            btn3 = types.KeyboardButton('Основное меню')
            markup.add(btn1, btn2)
            markup.add(btn3)
            bot.send_message(message.chat.id, 'Выберите действие', reply_markup = markup)

        elif message.text == 'Израсходовать':

            eq = change_BD('Расходные материалы')

            bot.send_message(message.chat.id, 'Израсходовать', reply_markup = main_menu_student())
            keyboard = types.InlineKeyboardMarkup()
            for i in range(len(eq)):
                name = change_BD('Расходные материалы')[i][:-1]
                button1 = (name +
                           f'(доступно: {str(exp_mat_kol(name, "Расходные материалы"))} {exp_mat_ed_izm(name, "Расходные материалы")})')
                keyboard.add(types.InlineKeyboardButton(text = button1, callback_data = 'spend_exp_mat' + name))
            bot.send_message(message.chat.id,'Список расходных материалов:', 
                             reply_markup = keyboard)    

        elif message.text == 'Вернуть':
            eq = change_BD('Расходные материалы')
            bot.send_message(message.chat.id, 'Вернуть', reply_markup = main_menu_student())
            keyboard = types.InlineKeyboardMarkup()
            for i in range(len(eq)):
                name = change_BD('Расходные материалы')[i][:-1]
                button2 = name
                
                keyboard.add(types.InlineKeyboardButton(text = button2, callback_data = 'cancel_spend_exp_mat' + name))
            bot.send_message(message.chat.id, 'Список расходных материалов:', 
                             reply_markup = keyboard)    


        # Вывод расписания кабинета
        elif message.text == 'Расписание кабинета':
            timetable(chat_id)

        # Кнопка отмена для студентов
        elif message.text == 'Отмена':
            bot.send_message(chat_id, 'Отменено', reply_markup = keyboard_main_menu_student())

        # Кнопка основное меню для студентов
        elif message.text == 'Основное меню':
            bot.send_message(chat_id, 'Возвращено в основное меню', reply_markup = keyboard_main_menu_student())

        else:
            bot.send_message(chat_id, 'Я вас не понимаю')
            bot.send_message(chat_id, 'Выберите пункт из меню', reply_markup = keyboard_main_menu_student())





#--------ФУНКЦИИ---------#  
#==========Считывание данных из базы данных============#

# Открываем BD (возвращаем страницу)
def read_BD(kategory):
    #wb = load_workbook(config.BD)
    wb = load_workbook('BD.xlsx')
    sheet = wb[kategory]
    return sheet 

 # Вывод списка элементов категории (возвращает первый столбец)
def change_BD(kategory):
    sheet = read_BD(kategory)

    data = []
    i=2
    val = ''
    while(True):
        val = sheet.cell(row = i, column = 1).value
        if val == None:
            break
        data.append(val + '\n')
        i+=1
    return data

#==========Регистрация пользователей============#

# Проверка начал ли пользователь регистрацию
def check_start_reg(chat_id, send_msg=False):
    if chat_id in BUFFER:
        if BUFFER[chat_id] == 'студент' or BUFFER[chat_id] == 'посетитель':
            data = BUFFER[chat_id]
            markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
            btn = types.KeyboardButton('Начать регистрацию с начала') 
            markup.add(btn)
            if send_msg:
                msg = bot.send_message(chat_id, 'Вы уже начали регистрацию как ' + data + 
                            '\nЕсли хотите начать регистрацию с начала, нажмите на кнопку Начать регистрацию с начала\n'+
                           'Если хотите продолжить регистрацию как ' + data +
                          ', продолжайте вводить данные', reply_markup = markup)

            return True 
    return False
           
# Удаляем данные из буффера и бд для пользователей, если пользователь нажал Начать регистрацию с начала             
def repeat_reg(message):
    if message.text == 'Начать регистрацию с начала':
        data = read()
        if str(message.chat.id) in read():
            del data[str(message.chat.id)]
            save(data)
        global BUFFER

        if message.chat.id in BUFFER:
            del BUFFER[message.chat.id]
        return True
    return False

    # Добавляем пользователя
def add_user(chat_id, data='', key=''): 
    chat_id = str(chat_id)
    
    content = read()
    if key == 'admin':
        s = {'status': 'admin'}
        content[chat_id] = s
        save(content)
        return

    elif key:

        if key == 'group':
            content[chat_id]['group'] = data
        else:
            s = {'status': 'student'}
            s[key] = data
            content[chat_id] = s
    
    else:
        s = {'status': 'visitor'}
        content[chat_id] = s
        content[chat_id]['number'] = data
    save(content) 
    return 


# Считываем список пользователей
def read(): 
    if os.path.isfile('users.json'): 
        file=open('users.json').read() 
        if file: 
            content=json.loads(file) 
            return content 
    return {} 

# Сохраняем список пользователей
def save(content): 
    file=open('users.json', 'wt') 
    file.write(json.dumps(content))
#====================================================================================================#

#==========================================ДЛЯ АДМИНИСТРАТОРОВ=======================================#

def keyboard_main_menu_admin():
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=False, resize_keyboard=True)
    btn1 = types.KeyboardButton('Написать сообщение посетителям')
    btn2 = types.KeyboardButton('Написать сообщение студентам')
    markup.add(btn1)
    markup.add(btn2)
    return markup

def send_message_visitors(message):
    text = message.text
    if text=='Отмена':
        return bot.send_message(message.chat.id, 'Отменено', reply_markup = keyboard_main_menu_admin())
    users = read()
    for i in users.keys():
        if users[i]['status'] == 'visitor':
            bot.send_message(i, text)
    bot.send_message(message.chat.id, 'Сообщение отправлено', reply_markup = keyboard_main_menu_admin())

def send_message_students(message):
    text = message.text
    if text=='Отмена':
        return bot.send_message(message.chat.id, 'Отменено', reply_markup = keyboard_main_menu_admin())
    users = read()
    for i in users.keys():
        if users[i]['status'] == 'student':
            bot.send_message(i, text)
    bot.send_message(message.chat.id, 'Сообщение отправлено', reply_markup = keyboard_main_menu_admin())
            

    

#====================================================================================================#




#=============================================ДЛЯ ПОСЕТИТЕЛЕЙ========================================#

# Задолжность
def owe(chat_id):
    sheet = read_BD('Информация для посетителей')
    i = search_contract_number(chat_id)
    j = search_categories('Задолженность по оплате')

    sum = sheet.cell(row = i, column = j).value
    if sum == None:
        return 'Задолжности нет'
    return f'Задолжность по оплате {sum} руб'


# Ближайшая отработка
def otrabotka(chat_id):
    sheet = read_BD('Информация для посетителей')
    i = search_contract_number(chat_id)
    j = search_categories('Ближайшая отработка')

    date = str(sheet.cell(row = i, column = j).value)
    time_lesson = str(sheet.cell(row = i, column = j+1).value)[:-3] 
    return timer.timer_otrabotka(date, time_lesson)


# Расписание занятий
def timetable_visitor(chat_id):
    sheet = read_BD('Информация для посетителей')
    i = search_contract_number(chat_id)
    j = search_categories('Занятие 1')

    weekday_lesson1 = sheet.cell(row = i, column = j).value
    time_lesson1 = (str(sheet.cell(row = i, column = j+1).value))[:-3]
    weekday_lesson2 = sheet.cell(row = i, column = j+2).value
    time_lesson2 = (str(sheet.cell(row = i, column = j+3).value))[:-3]
    cabinet = sheet.cell(row = i, column = j+4).value
    teacher_name = sheet.cell(row = i, column = j+5).value

    msg = ('Расписание занятий:\n'+
           f'{weekday_lesson1}  {time_lesson1}\n' + 
           f'{weekday_lesson2}  {time_lesson2}\n' +
           f'Аудитория {cabinet}\n\n' +
           f'Преподаватель\n{teacher_name}\n\n'+
           timer.timer_lesson(weekday_lesson1, time_lesson1, weekday_lesson2, time_lesson2))
    return msg


# Номер столбца
def search_categories(name): # принимает название столбца
    sheet = read_BD('Информация для посетителей')
    j=1
    while True:
        val = sheet.cell(row = 1, column = j).value
        if  str(val) == None:
            return print('Ошибка! Категория в БД не найдена')
        if  str(val) == name:
            return j 
            break
        j+=1

# Номер строки которая соответствует номеру договора
def search_contract_number(chat_id):
    sheet = read_BD('Информация для посетителей')
    i=3
    j = search_categories('Номер договора')

    number = read()[str(chat_id)]['number']
    while(True):
        val = sheet.cell(row = i, column = j).value

        if number == str(val):
            return i

        if sheet.cell(row = i, column = j).value == None :
            #bot.send_message(chat_id, 'Произошла ошибка. Попробуйте снова через какое-то время. Если это потвориться обратитесь к преподавателю') 
            return False
        i+=1

#==========Регистрация пользователей (посетитель)============#

# Ввод номера договора (для посетителей)
def ask_contract(message):

    if repeat_reg(message):
        return registration(message)

    chat_id = message.chat.id
    text = message.text

    if not text.isdigit():
        msg = bot.send_message(chat_id, 'Номер договора должен состоять из цифр, введите ещё раз.')
        bot.register_next_step_handler(msg, ask_contract)
        return

    add_user(chat_id, text)
    

    # Проверка ФИО посетителя
    sheet = read_BD('Информация для посетителей')
    i = search_contract_number(chat_id)
    j = search_categories('ФИО посетителя')

    if not i:
        content = read()
        del content[str(chat_id)]
        save(content)
        msg = bot.send_message(chat_id, 'Номер вашего договора отсутстует в списке! ' + 
                               'Введите номер договора повторно, если ситуация повториться, обратитесь к преподавателю')
        return bot.register_next_step_handler(msg, ask_contract)
    
    

    user_name = sheet.cell(row = i, column = j).value

    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    btn1 = types.KeyboardButton('Да')
    btn2 = types.KeyboardButton('Нет')
    markup.add(btn1, btn2)

    msg = bot.send_message(chat_id, f'ФИО посетителя {user_name}?', reply_markup = markup)
    bot.register_next_step_handler(msg, name_confirmation)


# Подтверждение правильности введенных данных для посетителей
def name_confirmation(message):
    chat_id = message.chat.id
    if message.text == 'Нет':
        content = read()
        del content[str(chat_id)]
        save(content)
        msg = bot.send_message(chat_id, 'Введите номер договора повторно, если ситуация повториться, обратитесь к преподавателю')
        bot.register_next_step_handler(msg, ask_contract)
    elif message.text == 'Да':
        del BUFFER[chat_id]
        bot.send_message(chat_id, 'Вы успешно зарегистрировались!', reply_markup = keyboard_main_menu_visitor())

# Основное меню для посетителей
def keyboard_main_menu_visitor():
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=False, resize_keyboard=True)
    btn1 = types.KeyboardButton('Расписание занятий')
    btn2 = types.KeyboardButton('Задолжность по оплате')
    btn3 = types.KeyboardButton('Ближайшая отработка')
    markup.add(btn1)
    markup.add(btn2, btn3)
    return markup


#====================================================================================================#








#======================================================ДЛЯ СТУДЕНТОВ==============================================#

#==========Расходные материалы============#

# Возвращает количество расходного материала
def exp_mat_kol(name, kategory):
      sheet = read_BD(kategory)
      i=2
      while(True):
        
        val = sheet.cell(row = i, column = 1).value

        if str(name) == str(val):
            if sheet.cell(row = i, column = 2).value == None :
                return '0'
            else:
                return (sheet.cell(row = i, column = 2).value)
        i+=1

# Возвращает eдeницы измерения расходного материала
def exp_mat_ed_izm(name, kategory):
      sheet = read_BD(kategory)
      i=2
      while(True):
        
        val = sheet.cell(row = i, column = 1).value
        if str(name) == str(val):
            if sheet.cell(row = i, column = 2).value == None :
                return ''
            else:
                return (sheet.cell(row = i, column = 3).value)
        i+=1

 # Изменить кол-во рас мат в базе данных 
def change_kol(message): 
    chat_id = message.chat.id
    text = message.text
    global BUFFER
    action = BUFFER[str(chat_id)][-1] # Действие которое нужно выполнить(s-spend, b-back)

    name = BUFFER[str(chat_id)][:-1] # Название материала
    message_id = BUFFER['m_id' + str(chat_id)] # id сообщения для редактирования количества

    if text == 'Отмена':

        bot.send_message(chat_id, 'Отменено', reply_markup = main_menu_student())

        del BUFFER[str(chat_id)]
        del BUFFER['m_id' + str(chat_id)]
        return 

    try:
        data = float(text.replace(',', '.'))
    except:
        msg = bot.send_message(chat_id, 'Введите число!')
        return bot.register_next_step_handler(msg, change_kol)

    if data <= 0: # Если пользователь ввел значение меньше нуля
        msg = bot.send_message(chat_id, 'Значение должно быть больше нуля!')
        return bot.register_next_step_handler(msg, change_kol)



    wb = load_workbook('BD.xlsx')
    sheet = wb['Расходные материалы']
    value = float(exp_mat_kol(name, 'Расходные материалы'))
    last_value = value

    if int(last_value) == float(last_value):
        last_value = int(last_value)


    if action == 's':
        value -= data
        if value < 0: # Если пользователь расходует больше чем доступно
            msg = bot.send_message(chat_id, f'Такого количества расходного материала нет (доступно: {last_value})')
            return bot.register_next_step_handler(msg, change_kol)
    elif action == 'b':
        value += data

   # записывем новое значение в БД
    i=2
    while(True):
       val = sheet.cell(row = i, column = 1).value
       if str(name) == str(val):
           sheet.cell(row = i, column = 2).value = value
           wb.save('BD.xlsx')
           break
       i+=1




    if action == 's': 
        # Меняем текст сообщения (количество)
        eq = change_BD('Расходные материалы')
        keyboard = types.InlineKeyboardMarkup()
        for i in range(len(eq)):
            name = change_BD('Расходные материалы')[i][:-1]
            button1 = (name +
                        f'(доступно: {str(exp_mat_kol(name, "Расходные материалы"))} {exp_mat_ed_izm(name, "Расходные материалы")})')
            keyboard.add(types.InlineKeyboardButton(text = button1, callback_data = 'spend_exp_mat' + name))
        bot.edit_message_text('Список расходных материалов:', chat_id, message_id, reply_markup = keyboard)


        bot.send_message(chat_id, f'Израсходовано {message.text} {exp_mat_ed_izm(name, "Расходные материалы")}',
                         reply_markup = main_menu_student())
    if action == 'b': 
        bot.send_message(chat_id, f'Возвращено {message.text} {exp_mat_ed_izm(name, "Расходные материалы")}', 
                         reply_markup = main_menu_student())
    del BUFFER[str(chat_id)]
    del BUFFER['m_id' + str(chat_id)]
    return

#==========Оборудование============#

# Возвращает описание оборудования
def info_equipment(name, kategory): 
      sheet = read_BD(kategory)
      i=2
      while(True):
        
        val = sheet.cell(row = i, column = 1).value

        if str(name) == str(val):
            if sheet.cell(row = i, column = 2).value == None :
                return '\n\nОписание не найдено'
            else:
                return '\n\nОписание:\n' + sheet.cell(row = i, column = 2).value
        i+=1



#===========Расписание кабинета===========#
def timetable(chat_id):
    sheet = read_BD('Расписание кабинета')
 
    i=2
    data = ''
    while(True):
        day = sheet.cell(row = i, column = 1).value        
        if day == None:
            return bot.send_message(chat_id, data, reply_markup = main_menu_student())
        time = sheet.cell(row = i, column = 2).value

        data = data + f'{day}    {time}\n'
        
        i+=1



#==========Регистрация пользователей (студент)============#

# Ввод ФИО (для студентов)
def ask_name(message):

    if repeat_reg(message):
        return registration(message)

    chat_id = message.chat.id
    text = message.text

    if text == 'Начать регистрацию с начала':
       return registration(message)

    if text.replace(' ', '').isdigit():
        msg = bot.send_message(chat_id, 'ФИО не может состоять из цифр, введите ещё раз.')
        bot.register_next_step_handler(msg, ask_name)
        return

    add_user(chat_id, text, 'name')
    msg = bot.send_message(chat_id, 'Введите группу (в формате XXX-16-1)')
    bot.register_next_step_handler(msg, ask_group)

# Ввод группы (для студентов)
def ask_group(message):

    if repeat_reg(message):
        return registration(message)

    chat_id = message.chat.id
    text = message.text

    add_user(chat_id, text, 'group')
    
    bot.send_message(chat_id, 'Вы успешно зарегистрировались!', reply_markup = keyboard_main_menu_student())
    del BUFFER[chat_id]

# Основное меню студентов 
def keyboard_main_menu_student():
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=False, resize_keyboard=True)
    btn1 = types.KeyboardButton('Оборудование')
    btn2 = types.KeyboardButton('Инструмент')
    btn3 = types.KeyboardButton('Расходные материалы')
    btn4 = types.KeyboardButton('Расписание кабинета')
    markup.add(btn1, btn2, btn3, btn4)
    return markup

# Формирует кнопку Основное меню для студентов     
def main_menu_student():
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=False, resize_keyboard=True)
    btn1 = types.KeyboardButton('Основное меню')
    markup.add(btn1)
    return markup
#====================================================================================================#



print('Бот запущен...')
bot.polling()

